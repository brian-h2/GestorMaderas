from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
from openpyxl import load_workbook
import io, os, json
from openpyxl.drawing.image import Image as XLImage


app = Flask(__name__)

# En producción (Railway/Render) seteá ALLOWED_ORIGIN con tu URL de Vercel
allowed_origin = os.environ.get("ALLOWED_ORIGIN", "*")
CORS(app, origins=allowed_origin)


@app.route("/", methods=["GET"])
def health():
    return jsonify({"status": "ok", "service": "Maderas Caroya API"})


@app.route("/insertar", methods=["POST"])
def insertar():
    filas_json = request.form.get("filas", "[]")

    try:
        filas = json.loads(filas_json)
    except json.JSONDecodeError:
        return jsonify({"error": "JSON de filas inválido"}), 400

    plantilla_path = os.path.join(os.path.dirname(__file__), "plantilla.xlsx")
    
    try:
        if os.path.exists(plantilla_path):
            wb = load_workbook(filename=plantilla_path)
            ws = wb.active
        else:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Pedidos"
            # Set up basic headers if creating from scratch
            headers = ["Nombre Pieza", "Material", "Largo", "Ancho", "Cantidad", "Veta", "Canto L. Sup", "Canto L. Inf", "Canto Izq.", "Canto Der.", "Notas"]
            for col, header in enumerate(headers, start=1):
                ws.cell(row=10, column=col, value=header)
    except Exception as e:
        return jsonify({"error": f"No se pudo inicializar el Excel: {str(e)}"}), 500

    if len(ws._images) == 0:
        logo_path = os.path.join(os.path.dirname(__file__), "logo.png")
        if os.path.exists(logo_path):
            try:
                img = XLImage(logo_path)
                img.anchor = "A1"  # ajustá la celda según tu plantilla
                ws.add_image(img)
            except Exception as e:
                print(f"Error al agregar la imagen: {e}")

    # Primera fila vacía a partir de la 11
    start_row = 11
    while ws.cell(row=start_row, column=1).value is not None:
        start_row += 1

    for i, f in enumerate(filas):
        r = start_row + i
        ws.cell(row=r, column=1).value  = f.get("nombre", "")
        ws.cell(row=r, column=2).value  = f.get("material", "")
        largo = f.get("largo", "")
        ws.cell(row=r, column=3).value  = float(largo) if largo else None
        ancho = f.get("ancho", "")
        ws.cell(row=r, column=4).value  = float(ancho) if ancho else None
        cant = f.get("cantidad", "1")
        ws.cell(row=r, column=5).value  = int(cant) if cant else 1
        ws.cell(row=r, column=6).value  = f.get("veta", "")
        ws.cell(row=r, column=7).value  = f.get("canto_largo_sup", "")
        ws.cell(row=r, column=8).value  = f.get("canto_largo_inf", "")
        ws.cell(row=r, column=9).value  = f.get("canto_izq", "")
        ws.cell(row=r, column=10).value = f.get("canto_der", "")
        ws.cell(row=r, column=11).value = f.get("notas", "")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    nombre_salida = "Plantilla_de_pedidos_Maderas_Caroya.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=nombre_salida,
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5055))
    app.run(host="0.0.0.0", port=port, debug=False)
